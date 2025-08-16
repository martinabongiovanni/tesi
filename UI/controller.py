import os
import subprocess
import sys
import openpyxl # per leggere, modificare e scrivere file Excel
from openpyxl import Workbook # per creare un nuovo file Excel
import flet as ft

class Controller:
    def __init__(self, view, model):
        self._view = view
        self._model = model
        self.lista_codice_area = [] # lista di tuple del tipo: (codice ateco, area) - (str, Business area)
        self._btn_apri_file = None

    def carica_dati_excel(self):
        '''
        Questa funzione legge i dati dal file in input, per ogni istanza preleva il codice ateco,
        chiama la funzione del model per analizzarlo e infine popola la lista_codice_area con gli
        elementi della lista_di_combinazioni di tuple generata nel model con le varie associazioni
        di codice ateco e business area.
        '''
        file_excel = openpyxl.load_workbook("database/AIDA_CoopB.xlsx")
        primo_sheet = file_excel.active
        # definisco una lista in cui ogni inserisco tutti gli elementi della prima riga del file
        intestazioni = [cella.value for cella in primo_sheet[1]]

        for row in primo_sheet.iter_rows(min_row=2, values_only=True):
            # definisco un dizionario con chiave l'intestazione della colonna e
            # valore il corrispettivo valore della riga in esame
            dati = dict(zip(intestazioni, row))

            # prelevo il codice ateco della coopB in esame
            codice = dati['ATECO 2002_codice']
            if codice is None:
                return
            # associo il codice a una business area chiamando il relativo metodo
            self._model.associazione_codiceAteco_BusinessArea(codice)

        # prelevo dal model la lista di combinazioni (codice, area)
        self.lista_codice_area = self._model.get_lista_combinazioni()

    def salva_excel_ordinato(self):
        '''
        Questa funzione serve per salvare i dati contenuti nella lista_codice_area in un nuovo file Excel.
        Saranno creati due sheet per ordinare i dati rispettivamente in ordine di codice (crescente) e
        in ordine di area (alfabetico).
        '''
        file_excel = Workbook()

        # 1. Ordinamento per codice
        primo_sheet_codice = file_excel.active
        primo_sheet_codice.title = "By Codice"
        self._scrivi_dati(primo_sheet_codice, sorted(self.lista_codice_area, key=lambda x: x[0]))

        # 2. Ordinamento per area
        secondo_sheet_area = file_excel.create_sheet("By Area")
        self._scrivi_dati(secondo_sheet_area, sorted(self.lista_codice_area, key=lambda x: x[1]))

        file_excel.save("output.xlsx")

    def _scrivi_dati(self, sheet, lista):
        '''
        Questa funzione scrive nello sheet che riceve come parametro i dati contenuti nella lista_codice_area
        in modo da avere nella prima colonna i codici ateco e nella seconda le aree con i relativi colori.
        :param sheet: Foglio del file excel da riempire.
        :param lista: La lista_codice_area che contiene tutte le tuple di associazioni (codice, area).
        '''
        # definisco le intestazioni
        sheet.append(["Codice Ateco", "Business Area"])

        # definisco le varie righe
        for elemento in lista:
            sheet.append([elemento[0], str(elemento[1])])
            colore = elemento[1].colora_area()
            sheet.cell(row=sheet.max_row, column=2).fill = colore

    def handle_ottieni_file(self, e):
        '''
        Alla pressione del bottone "Ottieni File", questa funzione genera il file con tutte le associazioni
        di codice ateco e business area.
        '''
        self.carica_dati_excel()
        self.salva_excel_ordinato()
        self._view.txt_result.controls.clear()
        row = ft.Row([ft.Text("Il file è stato generato correttamente!",
                                color="#202820", size=18, weight=ft.FontWeight.BOLD),
                      ft.Icon(ft.Icons.CHECK_CIRCLE_ROUNDED, color="#202820", size=30)])
        self._view.txt_result.controls.append(row)
        self._view.txt_result.controls.append(ft.Text('Clicca su "Apri" per visualizzarne il contento.',
                                                      color="#202820", size=16))
        self._btn_apri_file = ft.ElevatedButton(text="Apri", on_click=self.handle_apri_file,
                                              disabled=False, bgcolor="#faffeb", color="#202820")
        self._view.txt_result.controls.append(self._btn_apri_file)

        self._view.update_page()

    def handle_apri_file(self, e):
        '''
        Alla pressione del bottone "Apri", questa funzione permette la visualizzazione su Excel del file generato.
        '''
        file = "output.xlsx"

        if not os.path.exists(file):
            self._view.create_alert("Attenzione! File non trovato.")
            return

        try:
            if sys.platform.startswith('darwin'):  # macOS
                subprocess.call(('open', file))
            elif os.name == 'nt':  # Windows
                os.startfile(file)
            elif os.name == 'posix':  # Linux
                subprocess.call(('xdg-open', file))
            else:
                raise OSError("Sistema operativo non supportato.")

        finally:
            self._view.update_page()